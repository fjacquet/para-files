"""Tests for content extraction functions: Excel, ODS, ZIP, and 7Z readers.

All tests use tmp_path fixture with real file creation — no mocking of file format parsers.
"""

from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
from odf.opendocument import OpenDocumentSpreadsheet
from odf.table import Table, TableCell, TableRow
from odf.text import P

from para_files.utils.file_utils import (
    _read_archive_manifest,
    _read_excel_file,
    _read_ods_file,
    read_content_preview,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_xlsx(tmp_path: Path, sheet_name: str = "Sheet1", cell_value: str = "value") -> Path:
    """Create a minimal .xlsx file with a single sheet and one cell."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws["A1"] = cell_value
    path = tmp_path / "budget.xlsx"
    wb.save(str(path))
    return path


def _make_ods(tmp_path: Path, sheet_name: str = "Sheet1", cell_value: str = "value") -> Path:
    """Create a minimal .ods file with a single sheet and one cell."""
    doc = OpenDocumentSpreadsheet()
    table = Table(name=sheet_name)
    row = TableRow()
    cell = TableCell()
    cell.addElement(P(text=cell_value))
    row.addElement(cell)
    table.addElement(row)
    doc.spreadsheet.addElement(table)
    path = tmp_path / "plan.ods"
    doc.save(str(path))
    return path


def _make_zip(tmp_path: Path, filenames: list[str]) -> Path:
    """Create a .zip archive containing dummy entries for each filename."""
    path = tmp_path / "archive.zip"
    with zipfile.ZipFile(path, "w") as zf:
        for name in filenames:
            zf.writestr(name, b"dummy content")
    return path


# ---------------------------------------------------------------------------
# Excel (.xlsx) tests
# ---------------------------------------------------------------------------


class TestExcelXlsx:
    """Tests for _read_excel_file with .xlsx files."""

    def test_excel_xlsx_extracts_sheet_names(self, tmp_path: Path) -> None:
        """Sheet name 'Budget_2024' must appear in the extracted text."""
        path = _make_xlsx(tmp_path, sheet_name="Budget_2024", cell_value="Total Revenue")
        result = _read_excel_file(path, max_chars=2000)
        assert "Budget_2024" in result

    def test_excel_xlsx_extracts_cell_values(self, tmp_path: Path) -> None:
        """Cell value 'Total Revenue' must appear in the extracted text."""
        path = _make_xlsx(tmp_path, sheet_name="Budget_2024", cell_value="Total Revenue")
        result = _read_excel_file(path, max_chars=2000)
        assert "Total Revenue" in result

    def test_excel_dispatch_via_read_content_preview(self, tmp_path: Path) -> None:
        """read_content_preview() on .xlsx file must contain sheet/cell content."""
        path = _make_xlsx(tmp_path, sheet_name="Budget_2024", cell_value="Total Revenue")
        result = read_content_preview(path, max_chars=2000)
        assert "Budget_2024" in result or "Total Revenue" in result

    def test_excel_corrupt_returns_filename_fallback(self, tmp_path: Path) -> None:
        """A corrupt .xlsx file must return 'Filename: X' without raising an exception."""
        path = tmp_path / "bad.xlsx"
        path.write_bytes(b"not an xlsx file")
        result = _read_excel_file(path, max_chars=2000)
        assert result.startswith("Filename:")
        assert "bad.xlsx" in result

    def test_excel_max_chars_respected(self, tmp_path: Path) -> None:
        """Result length must not exceed max_chars even for large workbooks."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        # Fill many cells with data
        for row in range(1, 50):
            for col in range(1, 20):
                ws.cell(row=row, column=col, value=f"LongCellValue_row{row}_col{col}")
        path = tmp_path / "large.xlsx"
        wb.save(str(path))

        max_chars = 500
        result = _read_excel_file(path, max_chars=max_chars)
        assert len(result) <= max_chars


# ---------------------------------------------------------------------------
# ODS tests
# ---------------------------------------------------------------------------


class TestOds:
    """Tests for _read_ods_file with .ods files."""

    def test_ods_extracts_sheet_names(self, tmp_path: Path) -> None:
        """Sheet name 'Project Plan' must appear in the extracted text."""
        path = _make_ods(tmp_path, sheet_name="Project Plan", cell_value="Milestone 1")
        result = _read_ods_file(path, max_chars=2000)
        assert "Project Plan" in result

    def test_ods_extracts_cell_values(self, tmp_path: Path) -> None:
        """Cell value 'Milestone 1' must appear in the extracted text."""
        path = _make_ods(tmp_path, sheet_name="Project Plan", cell_value="Milestone 1")
        result = _read_ods_file(path, max_chars=2000)
        assert "Milestone 1" in result

    def test_ods_corrupt_returns_filename_fallback(self, tmp_path: Path) -> None:
        """A corrupt .ods file must return 'Filename: X' without raising an exception."""
        path = tmp_path / "bad.ods"
        path.write_bytes(b"random bytes")
        result = _read_ods_file(path, max_chars=2000)
        assert result.startswith("Filename:")
        assert "bad.ods" in result

    def test_ods_dispatch_via_read_content_preview(self, tmp_path: Path) -> None:
        """read_content_preview() on .ods file must return sheet/cell content."""
        path = _make_ods(tmp_path, sheet_name="Project Plan", cell_value="Milestone 1")
        result = read_content_preview(path, max_chars=2000)
        assert "Project Plan" in result or "Milestone 1" in result


# ---------------------------------------------------------------------------
# ZIP manifest tests
# ---------------------------------------------------------------------------


class TestZipManifest:
    """Tests for _read_archive_manifest with .zip files."""

    def test_zip_manifest_lists_filenames(self, tmp_path: Path) -> None:
        """Both 'invoice_2024.pdf' and 'contract.docx' must appear in the manifest."""
        path = _make_zip(tmp_path, ["invoice_2024.pdf", "contract.docx"])
        result = _read_archive_manifest(path, max_chars=2000)
        assert "invoice_2024.pdf" in result
        assert "contract.docx" in result

    def test_zip_manifest_format(self, tmp_path: Path) -> None:
        """Manifest must start with 'Archive manifest'."""
        path = _make_zip(tmp_path, ["file1.txt", "file2.txt"])
        result = _read_archive_manifest(path, max_chars=2000)
        assert result.startswith("Archive manifest")

    def test_zip_dispatch_via_read_content_preview(self, tmp_path: Path) -> None:
        """read_content_preview() on .zip must return the manifest text."""
        path = _make_zip(tmp_path, ["invoice_2024.pdf", "contract.docx"])
        result = read_content_preview(path, max_chars=2000)
        assert "Archive manifest" in result or "invoice_2024.pdf" in result

    def test_zip_corrupt_returns_filename_fallback(self, tmp_path: Path) -> None:
        """A corrupt .zip file must return 'Filename: X' without raising an exception."""
        path = tmp_path / "bad.zip"
        path.write_bytes(b"PK\x00garbage")
        result = _read_archive_manifest(path, max_chars=2000)
        assert result.startswith("Filename:")
        assert "bad.zip" in result

    def test_zip_max_chars_respected(self, tmp_path: Path) -> None:
        """Result length must not exceed max_chars even for archives with many entries."""
        filenames = [f"file_{i:04d}.txt" for i in range(500)]
        path = _make_zip(tmp_path, filenames)
        max_chars = 300
        result = _read_archive_manifest(path, max_chars=max_chars)
        assert len(result) <= max_chars


# ---------------------------------------------------------------------------
# Integration — read_content_preview dispatch
# ---------------------------------------------------------------------------


class TestReadContentPreviewDispatch:
    """Integration tests: verify read_content_preview routes to the correct extractor."""

    def test_xlsx_extension_dispatched(self, tmp_path: Path) -> None:
        """Extension .xlsx must route to the Excel reader, not return 'Filename: X'."""
        path = _make_xlsx(tmp_path, sheet_name="Budget_2024", cell_value="Total Revenue")
        result = read_content_preview(path, max_chars=2000)
        # A successful extraction means we got sheet/cell content — NOT a bare filename fallback
        assert result != f"Filename: {path.name}"
        assert "Budget_2024" in result or "Total Revenue" in result

    def test_ods_extension_dispatched(self, tmp_path: Path) -> None:
        """Extension .ods must route to the ODS reader, not return 'Filename: X'."""
        path = _make_ods(tmp_path, sheet_name="Project Plan", cell_value="Milestone 1")
        result = read_content_preview(path, max_chars=2000)
        assert result != f"Filename: {path.name}"
        assert "Project Plan" in result or "Milestone 1" in result

    def test_zip_extension_dispatched(self, tmp_path: Path) -> None:
        """Extension .zip must route to the archive reader, not return 'Filename: X'."""
        path = _make_zip(tmp_path, ["invoice_2024.pdf", "contract.docx"])
        result = read_content_preview(path, max_chars=2000)
        assert result != f"Filename: {path.name}"
        assert "invoice_2024.pdf" in result or "Archive manifest" in result
