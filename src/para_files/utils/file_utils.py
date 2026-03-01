"""File utility functions for metadata extraction and content reading."""

from __future__ import annotations

from collections.abc import Callable
from datetime import UTC, datetime
from pathlib import Path
from typing import TYPE_CHECKING

from loguru import logger

from para_files.types import FileMetadata


if TYPE_CHECKING:
    from para_files.utils.ocr import OCRResult
    from para_files.utils.pandoc import PandocResult


# Text file extensions that can be read directly
TEXT_EXTENSIONS = frozenset(
    {
        ".txt",
        ".md",
        ".markdown",
        ".rst",
        ".py",
        ".js",
        ".ts",
        ".json",
        ".yaml",
        ".yml",
        ".toml",
        ".xml",
        ".html",
        ".htm",
        ".css",
        ".csv",
        ".log",
        ".ini",
        ".cfg",
        ".conf",
        ".sh",
        ".bash",
        ".zsh",
        ".fish",
        ".sql",
        ".r",
        ".rb",
        ".php",
        ".java",
        ".c",
        ".cpp",
        ".h",
        ".hpp",
        ".go",
        ".rs",
        ".swift",
        ".kt",
        ".scala",
        ".clj",
        ".edn",
        ".ex",
        ".exs",
        ".erl",
        ".hrl",
        ".hs",
        ".ml",
        ".mli",
        ".elm",
        ".vue",
        ".svelte",
    }
)


# Spreadsheet extensions with extractable content
EXCEL_EXTENSIONS = frozenset({".xlsx", ".xlsm", ".xls"})

# Archive extensions with manifest-readable content
ARCHIVE_EXTENSIONS = frozenset({".zip", ".7z", ".7zip"})

# Maximum cells to extract per spreadsheet (prevents runaway on huge sheets)
_SPREADSHEET_MAX_CELLS = 200


# Media-only extensions that should NEVER be OCR'd for classification
# These are pure photo/video files - their content (if any text) is irrelevant
# OCR'ing these files can lead to misclassification (e.g., photo of certificate → certifications)
MEDIA_ONLY_EXTENSIONS = frozenset(
    {
        # Camera raw formats
        ".heic",  # Apple HEIC photos
        ".heif",  # HEIF photos
        ".raw",
        ".cr2",  # Canon RAW
        ".cr3",  # Canon RAW 3
        ".nef",  # Nikon RAW
        ".arw",  # Sony RAW
        ".dng",  # Adobe DNG
        ".orf",  # Olympus RAW
        ".rw2",  # Panasonic RAW
        # Video formats
        ".mp4",
        ".mov",
        ".avi",
        ".mkv",
        ".m4v",
        ".webm",
        ".wmv",
        ".flv",
        # Audio (no OCR possible but for completeness)
        ".mp3",
        ".m4a",
        ".wav",
        ".flac",
        ".aac",
        ".ogg",
    }
)


def extract_file_metadata(file_path: Path, *, extract_exif: bool = True) -> FileMetadata:
    """Extract metadata from a file.

    Args:
        file_path: Path to the file.
        extract_exif: Whether to extract EXIF data (requires exiftool).

    Returns:
        FileMetadata with filename, extension, size, dates, and optional EXIF/PDF.
    """
    stat = file_path.stat()

    # Get modification time as datetime
    modified_at = datetime.fromtimestamp(stat.st_mtime, tz=UTC)

    # Get creation time if available (macOS)
    created_at = None
    if hasattr(stat, "st_birthtime"):
        created_at = datetime.fromtimestamp(stat.st_birthtime, tz=UTC)

    # Extract EXIF data if requested
    exif_date = None
    exif_gps_lat = None
    exif_gps_lon = None
    exif_camera = None

    if extract_exif:
        from para_files.utils.exiftool import extract_exif as get_exif

        exif = get_exif(file_path)
        if exif:
            exif_date = exif.best_date
            if exif.gps:
                exif_gps_lat = exif.gps.latitude
                exif_gps_lon = exif.gps.longitude
            if exif.make and exif.model:
                exif_camera = f"{exif.make} {exif.model}"
            elif exif.make:
                exif_camera = exif.make
            elif exif.model:
                exif_camera = exif.model

    # Extract PDF metadata for PDF files
    pdf_title = None
    pdf_author = None
    pdf_subject = None

    if file_path.suffix.lower() == ".pdf":
        from para_files.utils.pdf_metadata import extract_pdf_metadata

        pdf_meta = extract_pdf_metadata(file_path, max_pages_for_isbn=0)  # Skip ISBN scan
        if pdf_meta:
            pdf_title = pdf_meta.title
            pdf_author = pdf_meta.author
            pdf_subject = pdf_meta.subject

    return FileMetadata(
        path=file_path,
        filename=file_path.name,
        extension=file_path.suffix.lower(),
        size_bytes=stat.st_size,
        modified_at=modified_at,
        created_at=created_at,
        exif_date=exif_date,
        exif_gps_lat=exif_gps_lat,
        exif_gps_lon=exif_gps_lon,
        exif_camera=exif_camera,
        pdf_title=pdf_title,
        pdf_author=pdf_author,
        pdf_subject=pdf_subject,
    )


def read_content_preview(
    file_path: Path,
    max_chars: int = 2000,
) -> str:
    """Read content preview from a file.

    For text files, reads the first max_chars characters.
    For documents (docx, odt, epub, etc.), uses pandoc to extract text.
    For PDF files, uses pypdf.
    For binary files, returns the filename as content.

    Args:
        file_path: Path to the file.
        max_chars: Maximum characters to read.

    Returns:
        Content preview string.
    """
    extension = file_path.suffix.lower()

    # For text files, read content directly
    if extension in TEXT_EXTENSIONS:
        return _read_text_file(file_path, max_chars)

    # For PDF files, try to extract text with pypdf
    if extension == ".pdf":
        return _read_pdf_file(file_path, max_chars)

    # For Excel spreadsheets, extract sheet names and cell values
    if extension in EXCEL_EXTENSIONS:
        return _read_excel_file(file_path, max_chars)

    # For ODS spreadsheets, extract sheet names and cell values
    if extension == ".ods":
        return _read_ods_file(file_path, max_chars)

    # For archive files, read the internal filename manifest (no extraction)
    if extension in ARCHIVE_EXTENSIONS:
        return _read_archive_manifest(file_path, max_chars)

    # For document formats, try pandoc extraction
    from para_files.utils.pandoc import PANDOC_EXTENSIONS, extract_text

    if extension in PANDOC_EXTENSIONS:
        return _read_document_file(file_path, max_chars, extract_text)

    # DEFENSE-IN-DEPTH: Skip OCR for pure media files (photos, videos, audio)
    # Even if they technically support OCR, extracting text from photos can cause
    # misclassification (e.g., a photo of a certificate → routed to certifications)
    if extension in MEDIA_ONLY_EXTENSIONS:
        logger.debug("Skipping OCR for media file: {}", file_path)
        return f"Filename: {file_path.name}"

    # For image files that ARE NOT pure media (e.g., scanned documents), try OCR
    from para_files.utils.ocr import OCR_EXTENSIONS
    from para_files.utils.ocr import extract_text as extract_ocr_text

    if extension in OCR_EXTENSIONS:
        return _read_image_file(file_path, max_chars, extract_ocr_text)

    # For other files, use filename as content
    return f"Filename: {file_path.name}"


def _read_text_file(file_path: Path, max_chars: int) -> str:
    """Read text file content.

    Args:
        file_path: Path to text file.
        max_chars: Maximum characters to read.

    Returns:
        Text content.
    """
    try:
        with file_path.open(encoding="utf-8", errors="replace") as f:
            return f.read(max_chars)
    except OSError:
        logger.warning("Failed to read text file: {}", file_path)
        return f"Filename: {file_path.name}"


# Minimum chars to consider PDF as having extractable text (vs scanned)
_PDF_MIN_TEXT_THRESHOLD = 50

# Minimum quality score to consider text as useful (0.0 - 1.0)
_PDF_MIN_QUALITY_THRESHOLD = 0.3

# Text quality evaluation thresholds
_METADATA_HIT_THRESHOLD = 3  # Number of metadata patterns to trigger rejection
_SPACE_RATIO_MIN = 0.08  # Minimum space ratio for good text
_SPACE_RATIO_MAX = 0.30  # Maximum space ratio for good text
_WORD_LEN_MIN = 3  # Minimum average word length
_WORD_LEN_MAX = 12  # Maximum average word length
_MIN_WORD_LEN = 2  # Minimum length to consider a word
_WORD_ALPHA_RATIO = 0.6  # Minimum alphabetic ratio for "real" words
_SHORT_TEXT_THRESHOLD = 100  # Text below this gets penalty
_MEDIUM_TEXT_THRESHOLD = 200  # Text below this gets smaller penalty

# Common PDF metadata patterns that indicate garbage text
_PDF_METADATA_PATTERNS = frozenset(
    {
        "created by",
        "pdf producer",
        "pdf version",
        "endobj",
        "startxref",
        "%%eof",
        "/type",
        "/page",
        "obj<<",
        "stream",
        "endstream",
        "xref",
        "/font",
        "/encoding",
    }
)


def _calculate_text_quality(text: str) -> float:
    """Calculate quality score for extracted text.

    Evaluates if text is useful content vs garbage/metadata.

    Args:
        text: Extracted text to evaluate.

    Returns:
        Quality score between 0.0 (garbage) and 1.0 (excellent).
    """
    if not text or not text.strip():
        return 0.0

    text = text.strip()
    text_lower = text.lower()

    # Check for PDF metadata patterns (strong negative signal)
    metadata_hits = sum(1 for p in _PDF_METADATA_PATTERNS if p in text_lower)
    if metadata_hits >= _METADATA_HIT_THRESHOLD:
        logger.debug("Text contains {} PDF metadata patterns", metadata_hits)
        return 0.0

    # Calculate character composition metrics
    total_chars = len(text)
    alpha_chars = sum(1 for c in text if c.isalpha())
    digit_chars = sum(1 for c in text if c.isdigit())
    space_chars = sum(1 for c in text if c.isspace())
    alnum_chars = alpha_chars + digit_chars

    # Ratio of alphanumeric characters (should be high for real text)
    alnum_ratio = alnum_chars / total_chars if total_chars > 0 else 0

    # Ratio of alphabetic characters (real text has mostly letters)
    alpha_ratio = alpha_chars / total_chars if total_chars > 0 else 0

    # Space ratio (real text has ~15-20% spaces, garbage often has very few or too many)
    space_ratio = space_chars / total_chars if total_chars > 0 else 0
    space_score = 1.0 if _SPACE_RATIO_MIN <= space_ratio <= _SPACE_RATIO_MAX else 0.5

    # Word analysis - split on whitespace and check word quality
    words = text.split()
    if not words:
        return 0.0

    # Average word length (real words are typically 3-10 chars)
    avg_word_len = sum(len(w) for w in words) / len(words)
    word_len_score = 1.0 if _WORD_LEN_MIN <= avg_word_len <= _WORD_LEN_MAX else 0.5

    # Ratio of "real" words (mostly alphabetic, reasonable length)
    real_words = sum(
        1
        for w in words
        if len(w) >= _MIN_WORD_LEN and sum(1 for c in w if c.isalpha()) / len(w) > _WORD_ALPHA_RATIO
    )
    word_quality_ratio = real_words / len(words) if words else 0

    # Combine scores with weights
    quality_score = (
        alpha_ratio * 0.25  # Letters are important
        + alnum_ratio * 0.15  # Alphanumeric content
        + space_score * 0.20  # Proper spacing
        + word_len_score * 0.15  # Reasonable word lengths
        + word_quality_ratio * 0.25  # Quality of words
    )

    # Penalty for very short text (less confident)
    if total_chars < _SHORT_TEXT_THRESHOLD:
        quality_score *= 0.7
    elif total_chars < _MEDIUM_TEXT_THRESHOLD:
        quality_score *= 0.85

    logger.debug(
        "Text quality: {:.2f} (alpha={:.2f}, alnum={:.2f}, space={:.2f}, "
        "word_len={:.1f}, word_qual={:.2f}, chars={})",
        quality_score,
        alpha_ratio,
        alnum_ratio,
        space_ratio,
        avg_word_len,
        word_quality_ratio,
        total_chars,
    )

    return min(1.0, max(0.0, quality_score))


def _should_try_ocr(text: str) -> bool:
    """Determine if OCR should be attempted based on text quality.

    Args:
        text: Text extracted by pypdf/pdftotext.

    Returns:
        True if OCR should be attempted.
    """
    stripped = text.strip()

    # Always try OCR if text is very short
    if len(stripped) < _PDF_MIN_TEXT_THRESHOLD:
        return True

    # Calculate quality and decide
    quality = _calculate_text_quality(stripped)

    if quality < _PDF_MIN_QUALITY_THRESHOLD:
        logger.debug(
            "Text quality {:.2f} below threshold {:.2f}, will try OCR",
            quality,
            _PDF_MIN_QUALITY_THRESHOLD,
        )
        return True

    return False


def _read_pdf_file(file_path: Path, max_chars: int) -> str:
    """Extract text from PDF file.

    Uses pypdf first, falls back to OCR for scanned PDFs or low-quality text.

    Args:
        file_path: Path to PDF file.
        max_chars: Maximum characters to extract.

    Returns:
        Extracted text or filename fallback.
    """
    # 1. Try pypdf first (fast, works for text-based PDFs)
    text = _extract_pdf_with_pypdf(file_path, max_chars)

    # 2. Check if text quality is sufficient, or if OCR should be attempted
    if _should_try_ocr(text):
        logger.debug(
            "PDF text quality insufficient, trying OCR: {}",
            file_path,
        )
        ocr_text = _ocr_pdf_first_page(file_path, max_chars)

        # Use OCR result if it's better quality
        if ocr_text:
            ocr_quality = _calculate_text_quality(ocr_text)
            text_quality = _calculate_text_quality(text)

            if ocr_quality > text_quality or len(ocr_text.strip()) > len(text.strip()):
                logger.debug(
                    "Using OCR text (quality {:.2f} vs {:.2f})",
                    ocr_quality,
                    text_quality,
                )
                return ocr_text

    return text if text.strip() else f"Filename: {file_path.name}"


def _extract_pdf_with_pypdf(file_path: Path, max_chars: int) -> str:
    """Extract text from PDF using pypdf, with pdftotext fallback."""
    # Try pypdf first
    try:
        from pypdf import PdfReader

        reader = PdfReader(file_path)
        text_parts = []
        chars_read = 0

        for page in reader.pages:
            page_text = page.extract_text() or ""
            remaining = max_chars - chars_read
            if remaining <= 0:
                break
            text_parts.append(page_text[:remaining])
            chars_read += len(page_text[:remaining])

        return "".join(text_parts)

    except ImportError:
        logger.debug("pypdf not installed: {}", file_path)
    except Exception:  # noqa: BLE001
        # pypdf can fail on various PDF formats - try pdftotext fallback
        logger.debug("pypdf failed, trying pdftotext: {}", file_path)

    # Fallback to pdftotext (poppler-utils)
    try:
        import subprocess

        result = subprocess.run(  # noqa: S603
            ["pdftotext", "-q", str(file_path), "-"],  # noqa: S607
            check=False,
            capture_output=True,
            text=True,
            timeout=30,
        )
        if result.returncode == 0 and result.stdout:
            return result.stdout[:max_chars]
    except (FileNotFoundError, subprocess.TimeoutExpired):
        logger.debug("pdftotext not available or timeout: {}", file_path)
    except Exception:  # noqa: BLE001
        logger.debug("pdftotext failed: {}", file_path)

    return ""


def _ocr_pdf_first_page(file_path: Path, max_chars: int) -> str:
    """OCR the first page of a PDF using pymupdf + Vision framework.

    Args:
        file_path: Path to PDF file.
        max_chars: Maximum characters to extract.

    Returns:
        OCR text or empty string on failure.
    """
    import tempfile

    try:
        import fitz  # pymupdf

        from para_files.utils.ocr import extract_text as ocr_extract
    except ImportError as e:
        logger.debug("pymupdf or OCR not available: {}", e)
        return ""

    try:
        # Open PDF and render first page to image
        doc = fitz.open(file_path)
        if len(doc) == 0:
            return ""

        page = doc[0]
        # Render at 2x resolution for better OCR
        mat = fitz.Matrix(2.0, 2.0)
        pix = page.get_pixmap(matrix=mat)

        # Convert to PNG bytes for Vision OCR
        png_bytes = pix.tobytes("png")

        # Create temp file for OCR (Vision needs a file path)
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp.write(png_bytes)
            tmp_path = Path(tmp.name)

        try:
            result = ocr_extract(tmp_path, max_chars)
        finally:
            tmp_path.unlink(missing_ok=True)

    except FileNotFoundError:
        raise
    except Exception:  # noqa: BLE001
        logger.exception("OCR failed for PDF: {}", file_path)
        return ""

    if result and result.text:
        logger.debug("OCR extracted {} chars from PDF: {}", len(result.text), file_path)
        return result.text[:max_chars]

    return ""


def _read_document_file(
    file_path: Path,
    max_chars: int,
    extract_fn: Callable[[Path, int], PandocResult | None],
) -> str:
    """Extract text from document file using pandoc.

    Args:
        file_path: Path to document file.
        max_chars: Maximum characters to extract.
        extract_fn: Function to extract text (injected to avoid circular import).

    Returns:
        Extracted text or filename fallback.
    """
    result = extract_fn(file_path, max_chars)
    if result and result.text:
        return result.text

    logger.debug("pandoc extraction failed, using filename: {}", file_path)
    return f"Filename: {file_path.name}"


def _read_image_file(
    file_path: Path,
    max_chars: int,
    extract_fn: Callable[[Path, int], OCRResult | None],
) -> str:
    """Extract text from image file using OCR.

    Args:
        file_path: Path to image file.
        max_chars: Maximum characters to extract.
        extract_fn: Function to extract text (injected to avoid circular import).

    Returns:
        Extracted text or filename fallback.
    """
    result = extract_fn(file_path, max_chars)
    if result and result.text:
        return result.text

    logger.debug("OCR extraction failed, using filename: {}", file_path)
    return f"Filename: {file_path.name}"


def _extract_xlsx_content(file_path: Path, max_chars: int) -> list[str]:
    """Extract text parts from an xlsx/xlsm workbook using openpyxl.

    Args:
        file_path: Path to xlsx/xlsm file.
        max_chars: Maximum total characters to collect.

    Returns:
        List of text parts (sheet names and cell values).
    """
    import openpyxl

    parts: list[str] = []
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        parts.append(f"Sheet: {sheet_name}")
        ws = wb[sheet_name]
        cell_count = 0
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None and str(cell).strip():
                    parts.append(str(cell).strip())
                    cell_count += 1
                    if cell_count >= _SPREADSHEET_MAX_CELLS:
                        break
            if cell_count >= _SPREADSHEET_MAX_CELLS:
                break
            if sum(len(p) for p in parts) >= max_chars:
                break
    wb.close()
    return parts


def _extract_xls_content(file_path: Path, max_chars: int) -> list[str]:
    """Extract text parts from a legacy .xls workbook using xlrd.

    Args:
        file_path: Path to xls file.
        max_chars: Maximum total characters to collect.

    Returns:
        List of text parts (sheet names and cell values).
    """
    import xlrd

    parts: list[str] = []
    wb = xlrd.open_workbook(str(file_path))
    for sheet_name in wb.sheet_names():
        parts.append(f"Sheet: {sheet_name}")
        ws = wb.sheet_by_name(sheet_name)
        cell_count = 0
        for row_idx in range(ws.nrows):
            for col_idx in range(ws.ncols):
                cell_val = ws.cell_value(row_idx, col_idx)
                if cell_val and str(cell_val).strip():
                    parts.append(str(cell_val).strip())
                    cell_count += 1
                    if cell_count >= _SPREADSHEET_MAX_CELLS:
                        break
            if cell_count >= _SPREADSHEET_MAX_CELLS:
                break
            if sum(len(p) for p in parts) >= max_chars:
                break
    return parts


def _read_excel_file(file_path: Path, max_chars: int) -> str:
    """Extract text from Excel file using openpyxl (xlsx/xlsm) or xlrd (xls).

    Reads sheet names and first N cell values from each sheet.
    Handles encrypted/corrupt files gracefully.

    Args:
        file_path: Path to Excel file.
        max_chars: Maximum characters to return.

    Returns:
        Extracted text or filename fallback.
    """
    extension = file_path.suffix.lower()
    parts: list[str] = []

    try:
        if extension in {".xlsx", ".xlsm"}:
            parts = _extract_xlsx_content(file_path, max_chars)
        elif extension == ".xls":
            # xlrd is only used for legacy .xls; prefer openpyxl for newer formats
            parts = _extract_xls_content(file_path, max_chars)

    except ImportError as e:
        logger.warning("Excel library not available for {}: {}", file_path, e)
        return f"Filename: {file_path.name}"
    except Exception:  # noqa: BLE001
        logger.warning("Failed to read Excel file (corrupt or encrypted?): {}", file_path)
        return f"Filename: {file_path.name}"

    text = "\n".join(parts)
    return text[:max_chars] if text.strip() else f"Filename: {file_path.name}"


def _extract_ods_content(file_path: Path, max_chars: int) -> list[str]:
    """Extract text parts from an ODS document using odfpy.

    Args:
        file_path: Path to ODS file.
        max_chars: Maximum total characters to collect.

    Returns:
        List of text parts (sheet names and cell values).
    """
    from odf.opendocument import load
    from odf.table import Table, TableCell, TableRow
    from odf.text import P

    doc = load(str(file_path))
    parts: list[str] = []
    cell_count = 0
    done = False

    for sheet in doc.spreadsheet.getElementsByType(Table):
        if done:
            break
        sheet_name = sheet.getAttribute("name") or "Sheet"
        parts.append(f"Sheet: {sheet_name}")

        for row in sheet.getElementsByType(TableRow):
            if done:
                break
            for cell in row.getElementsByType(TableCell):
                for p in cell.getElementsByType(P):
                    text = str(p)
                    if text.strip():
                        parts.append(text.strip())
                        cell_count += 1
                        if cell_count >= _SPREADSHEET_MAX_CELLS:
                            done = True
                            break
                if done or sum(len(p) for p in parts) >= max_chars:
                    done = True
                    break

    return parts


def _read_ods_file(file_path: Path, max_chars: int) -> str:
    """Extract text from ODS file using odfpy.

    Reads sheet names and first N cell values from each sheet.
    Handles corrupt files gracefully.

    Args:
        file_path: Path to ODS file.
        max_chars: Maximum characters to return.

    Returns:
        Extracted text or filename fallback.
    """
    try:
        parts = _extract_ods_content(file_path, max_chars)
    except ImportError as e:
        logger.warning("odfpy not available for {}: {}", file_path, e)
        return f"Filename: {file_path.name}"
    except Exception:  # noqa: BLE001
        logger.warning("Failed to read ODS file (corrupt?): {}", file_path)
        return f"Filename: {file_path.name}"

    text = "\n".join(parts)
    return text[:max_chars] if text.strip() else f"Filename: {file_path.name}"


def _read_archive_manifest(file_path: Path, max_chars: int) -> str:
    """Read archive manifest (list of internal filenames) without extracting.

    For ZIP files: uses stdlib zipfile.ZipFile.namelist() — no extraction.
    For 7Z files: uses py7zr.SevenZipFile.getnames() — no extraction.

    Handles encrypted, corrupt, and unsupported archives gracefully.

    Args:
        file_path: Path to archive file.
        max_chars: Maximum characters to return.

    Returns:
        Manifest text listing internal filenames, or filename fallback.
    """
    extension = file_path.suffix.lower()
    names: list[str] = []

    try:
        if extension == ".zip":
            import zipfile

            with zipfile.ZipFile(file_path, "r") as zf:
                names = zf.namelist()

        elif extension in {".7z", ".7zip"}:
            try:
                import py7zr

                with py7zr.SevenZipFile(file_path, mode="r") as zf:
                    names = zf.getnames()
            except ImportError:
                logger.debug("py7zr not installed, cannot read 7Z manifest: {}", file_path)
                return f"Filename: {file_path.name}"

    except Exception:  # noqa: BLE001
        logger.warning(
            "Failed to read archive manifest (corrupt, encrypted, or unsupported?): {}",
            file_path,
        )
        return f"Filename: {file_path.name}"

    if not names:
        return f"Filename: {file_path.name}"

    # Build manifest text: "Archive manifest:\n  invoice_2024.pdf\n  contract.docx\n..."
    header = f"Archive manifest ({len(names)} files):"
    parts = [header] + [f"  {name}" for name in names]
    text = "\n".join(parts)
    return text[:max_chars]
